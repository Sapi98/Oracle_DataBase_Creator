from openpyxl import load_workbook
import os


class CreateDataset:

    def __init__(self):
        cd = os.getcwd()
        self.path1 = cd + '/Indian_first_names.xlsx'
        self.path2 = cd + '/Indian_middle_names.xlsx'
        self.path3 = cd + '/Indian_last_names.xlsx'

    def generate_data_list(self, name_length=2):
        wb1 = load_workbook(filename=self.path1)
        wb3 = load_workbook(filename=self.path3)

        first_names = []
        last_names = []
        middle_names = []

        wb1_sheet = wb1.active
        wb3_sheet = wb3.active

        for i in range(wb1_sheet.max_row):
            first_names.append(wb1_sheet.cell(row=i+1, column=1))

        for i in range(wb3_sheet.max_row):
            last_names.append(wb3_sheet.cell(row=i+1, column=1))

        if name_length == 3:
            wb2 = load_workbook(filename=self.path2)
            wb2_sheet = wb2.active
            for i in range(wb2_sheet.max_row):
                middle_names.append(wb2_sheet.cell(row=i+1, column=1))

        result = (first_names, middle_names, last_names)

        return result
